"""Generic Skill Code Runner for instruction-only Agent Skills.

The runner asks the configured model to translate an activated SKILL.md plus the
user task into a short Python script, executes that script in a temporary
workspace, and persists files written to outputs/ into My Files.
"""

from __future__ import annotations

import ast
import base64
import json
import logging
import os
import re
import shutil
import subprocess
import sys
import tempfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from app.runtime.llm_service import LLMRequest, llm_service
from app.runtime.skill_service import get_skill_package_root
from app.services.user_file_service import user_file_service

logger = logging.getLogger(__name__)


_ALLOWED_IMPORT_ROOTS = {
    "base64",
    "csv",
    "datetime",
    "decimal",
    "io",
    "json",
    "math",
    "openpyxl",
    "os",
    "pandas",
    "pathlib",
    "re",
    "shutil",
    "statistics",
    "subprocess",
    "sys",
    "tempfile",
    "typing",
    "zipfile",
}

_FORBIDDEN_CALL_NAMES = {
    "eval",
    "exec",
    "compile",
    "input",
    "breakpoint",
    "__import__",
}

_FORBIDDEN_ATTR_NAMES = {
    "unlink",
    "rmdir",
    "remove",
    "removedirs",
    "rmtree",
    "rename",
    "replace",
}

_ARTIFACT_HINTS = {
    "file",
    "files",
    "download",
    "deliverable",
    "artifact",
    "workbook",
    "spreadsheet",
    "excel",
    "html",
    "report",
    "document",
    "presentation",
    "\u6587\u4ef6",
    "\u8868\u683c",
    "\u5de5\u4f5c\u7c3f",
    "\u9875\u9762",
    "\u7f51\u7ad9",
    "\u62a5\u8868",
    "\u4e0b\u8f7d",
    "\u4fdd\u5b58",
    "\u53e6\u5b58",
    "\u751f\u6210",
    "\u521b\u5efa",
    "\u5236\u4f5c",
    "\u4fee\u6539",
    "\u7f16\u8f91",
    "\u6e05\u7406",
    "\u8f6c\u6362",
    "\u5bfc\u51fa",
    ".xlsx",
    ".xlsm",
    ".xltx",
    ".csv",
    ".tsv",
    ".html",
    ".json",
    ".txt",
}


_BLANK_COLUMN_TERMS = {
    "blank column",
    "blank columns",
    "empty column",
    "empty columns",
    "\u7a7a\u767d\u5217",
    "\u7a7a\u5217",
}

_DELETE_TERMS = {
    "delete",
    "remove",
    "drop",
    "clean",
    "\u5220\u9664",
    "\u5220\u6389",
    "\u53bb\u6389",
    "\u6e05\u7406",
}

_SPREADSHEET_EXTENSIONS = {".xlsx", ".xlsm", ".xltx"}

@dataclass
class SkillCodeRunResult:
    handled: bool
    status: str = "skipped"
    answer: str = ""
    saved_files: list[str] = field(default_factory=list)
    error: str | None = None
    events: list[dict[str, Any]] = field(default_factory=list)


def should_attempt_skill_code_runner(skill: dict[str, Any], user_text: str, referenced_files: list[str]) -> bool:
    """Return True when an instruction-only Skill likely needs a file artifact."""
    if str(skill.get("entrypoint") or "").strip():
        return False

    text = (user_text or "").lower()
    if referenced_files:
        return True
    return any(hint in text for hint in _ARTIFACT_HINTS)


class SkillCodeRunner:
    def run(
        self,
        *,
        skill: dict[str, Any],
        user_text: str,
        user_id: str,
        referenced_files: list[str] | None = None,
        output_base_dir: str | None = None,
        model_provider: str | None = None,
        model_name: str | None = None,
        provider_profile: str | None = None,
        provider_connection_id: str | None = None,
        provider_connection: dict | None = None,
        timeout_seconds: int = 60,
    ) -> SkillCodeRunResult:
        referenced_files = list(referenced_files or [])
        if not should_attempt_skill_code_runner(skill, user_text, referenced_files):
            return SkillCodeRunResult(handled=False)

        skill_name = str(skill.get("name") or skill.get("skill_id") or "skill").strip() or "skill"
        safe_skill_name = re.sub(r"[^A-Za-z0-9_-]+", "_", skill_name).strip("_") or "skill"
        output_base_dir = output_base_dir or f"generated/{safe_skill_name}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"

        temp_root = Path(tempfile.mkdtemp(prefix="ha_skill_code_"))
        try:
            work_dir = temp_root / "work"
            inputs_dir = work_dir / "inputs"
            outputs_dir = work_dir / "outputs"
            skill_dir = work_dir / "skill"
            work_dir.mkdir(parents=True, exist_ok=True)
            inputs_dir.mkdir(parents=True, exist_ok=True)
            outputs_dir.mkdir(parents=True, exist_ok=True)

            package_root = get_skill_package_root(str(skill.get("skill_id") or ""))
            shutil.copytree(package_root, skill_dir, dirs_exist_ok=True)
            input_files = self._copy_input_files(user_id, referenced_files, inputs_dir)
            skill_files = self._list_skill_files(skill_dir)

            prompt = self._build_code_generation_prompt(
                skill=skill,
                user_text=user_text,
                input_files=input_files,
                skill_files=skill_files,
            )
            generation = llm_service.generate(
                LLMRequest(
                    text=prompt,
                    model_provider=model_provider,
                    model_name=model_name,
                    provider_profile=provider_profile,
                    provider_connection_id=provider_connection_id,
                    provider_connection=provider_connection,
                    max_tokens=6000,
                )
            )
            if not generation.ok:
                return SkillCodeRunResult(
                    handled=True,
                    status="failed",
                    error=generation.error or "model generation failed",
                    answer=(
                        f"已激活 {skill_name} Skill，但通用 Skill Code Runner 生成执行脚本失败。\n\n"
                        f"错误信息：{generation.error or 'model generation failed'}"
                    ),
                    events=[{
                        "stage": "skill_code_runner",
                        "status": "failed",
                        "message": "Failed to generate Skill execution code",
                        "payload": {"skill": skill_name, "error": generation.error},
                    }],
                )

            script_code = self._extract_script_code(generation.text)
            if not script_code.strip():
                return SkillCodeRunResult(
                    handled=True,
                    status="failed",
                    error="model did not return Python code",
                    answer=f"已激活 {skill_name} Skill，但模型没有返回可执行的 Python 脚本。",
                    events=[{
                        "stage": "skill_code_runner",
                        "status": "failed",
                        "message": "Generated response did not contain executable code",
                        "payload": {"skill": skill_name, "preview": (generation.text or "")[:1000]},
                    }],
                )

            try:
                self._validate_generated_code(script_code)
            except ValueError as exc:
                return SkillCodeRunResult(
                    handled=True,
                    status="failed",
                    error=str(exc),
                    answer=f"已激活 {skill_name} Skill，但生成的执行脚本未通过安全检查：{exc}",
                    events=[{
                        "stage": "skill_code_runner",
                        "status": "failed",
                        "message": "Generated Skill code failed validation",
                        "payload": {"skill": skill_name, "error": str(exc)},
                    }],
                )

            task_payload = {
                "user_text": user_text,
                "input_files": input_files,
                "inputs_dir": str(inputs_dir),
                "outputs_dir": str(outputs_dir),
                "skill_dir": str(skill_dir),
            }
            task_path = work_dir / "task.json"
            script_path = work_dir / "generated_skill_task.py"
            task_path.write_text(json.dumps(task_payload, ensure_ascii=False, indent=2), encoding="utf-8")
            script_path.write_text(self._wrap_script(script_code, task_path), encoding="utf-8")

            proc = subprocess.run(
                [sys.executable, str(script_path)],
                cwd=str(work_dir),
                text=True,
                capture_output=True,
                timeout=timeout_seconds,
                check=False,
            )
            stdout = (proc.stdout or "").strip()
            stderr = (proc.stderr or "").strip()
            if proc.returncode != 0:
                error = stderr or stdout or f"Skill code exited with {proc.returncode}"
                return SkillCodeRunResult(
                    handled=True,
                    status="failed",
                    error=error,
                    answer=f"已激活 {skill_name} Skill，但通用执行脚本运行失败。\n\n错误信息：{error}",
                    events=[{
                        "stage": "skill_code_runner",
                        "status": "failed",
                        "message": "Generated Skill code execution failed",
                        "payload": {"skill": skill_name, "stdout": stdout[-2000:], "stderr": stderr[-2000:]},
                    }],
                )

            generated_files = self._collect_output_files(outputs_dir)
            validation_issues = self._validate_generated_outputs(user_text, input_files, outputs_dir)
            if validation_issues:
                repair_message = self._repair_outputs_for_simple_postconditions(user_text, input_files, outputs_dir)
                generated_files = self._collect_output_files(outputs_dir)
                validation_issues = self._validate_generated_outputs(user_text, input_files, outputs_dir)
                if validation_issues:
                    issue_text = "; ".join(validation_issues)
                    return SkillCodeRunResult(
                        handled=True,
                        status="failed",
                        error=issue_text,
                        answer=f"已激活 {skill_name} Skill，但生成文件没有通过任务校验。\n\n校验问题：{issue_text}",
                        events=[{
                            "stage": "skill_code_runner",
                            "status": "failed",
                            "message": "Generated Skill output failed postcondition validation",
                            "payload": {"skill": skill_name, "issues": validation_issues, "repair_message": repair_message, "stdout": stdout[-2000:], "stderr": stderr[-2000:]},
                        }],
                    )
                if repair_message:
                    stdout = (stdout + "\n" + json.dumps({"ok": True, "message": repair_message}, ensure_ascii=False)).strip()
            if not generated_files:
                return SkillCodeRunResult(
                    handled=True,
                    status="failed",
                    error="no output files were created",
                    answer=f"已激活 {skill_name} Skill，但执行脚本没有在 outputs 目录生成文件。",
                    events=[{
                        "stage": "skill_code_runner",
                        "status": "failed",
                        "message": "Generated Skill code created no files",
                        "payload": {"skill": skill_name, "stdout": stdout[-2000:], "stderr": stderr[-2000:]},
                    }],
                )

            saved_paths = user_file_service.save_generated_files(user_id, output_base_dir, generated_files)
            message = self._extract_message(stdout) or f"已按请求使用 {skill_name} Skill 生成文件。"
            answer = (
                f"已使用 {skill_name} Skill 通过通用 Skill Code Runner 完成处理。\n\n"
                f"{message}\n\n"
                "文件已保存到 My Files：\n"
                + "\n".join(f"- {path}" for path in saved_paths)
                + "\n\n请到 My Files 中下载。"
            )
            return SkillCodeRunResult(
                handled=True,
                status="succeeded",
                answer=answer,
                saved_files=saved_paths,
                events=[{
                    "stage": "skill_code_runner",
                    "status": "succeeded",
                    "message": "Generated Skill code executed and saved files",
                    "payload": {"skill": skill_name, "files": saved_paths, "stdout": stdout[-2000:]},
                }],
            )
        except subprocess.TimeoutExpired:
            error = f"Skill code execution timeout after {timeout_seconds}s"
            return SkillCodeRunResult(
                handled=True,
                status="failed",
                error=error,
                answer=f"已激活 {skill_name} Skill，但通用执行脚本超时。\n\n错误信息：{error}",
                events=[{"stage": "skill_code_runner", "status": "failed", "message": error, "payload": {"skill": skill_name}}],
            )
        except Exception as exc:
            logger.error("Skill Code Runner failed", exc_info=True)
            return SkillCodeRunResult(
                handled=True,
                status="failed",
                error=str(exc),
                answer=f"已激活 {skill_name} Skill，但通用 Skill Code Runner 失败。\n\n错误信息：{exc}",
                events=[{"stage": "skill_code_runner", "status": "failed", "message": "Skill Code Runner failed", "payload": {"skill": skill_name, "error": str(exc)}}],
            )
        finally:
            shutil.rmtree(temp_root, ignore_errors=True)

    def _copy_input_files(self, user_id: str, referenced_files: list[str], inputs_dir: Path) -> list[dict[str, str]]:
        copied: list[dict[str, str]] = []
        used_names: set[str] = set()
        for index, relative_path in enumerate(referenced_files, start=1):
            try:
                source = user_file_service.get_file_for_download(user_id, relative_path)
                filename = source.name or f"input_{index}"
                if filename in used_names:
                    filename = f"{index}_{filename}"
                used_names.add(filename)
                target = inputs_dir / filename
                shutil.copy2(source, target)
                copied.append({
                    "source_path": relative_path,
                    "local_path": str(target),
                    "filename": filename,
                    "extension": target.suffix.lower(),
                })
            except Exception as exc:
                copied.append({"source_path": relative_path, "error": str(exc)})
        return copied

    def _list_skill_files(self, skill_dir: Path) -> list[str]:
        files: list[str] = []
        for path in skill_dir.rglob("*"):
            if not path.is_file():
                continue
            rel = path.relative_to(skill_dir).as_posix()
            if "__pycache__" in rel:
                continue
            files.append(rel)
            if len(files) >= 200:
                break
        return files

    def _build_code_generation_prompt(
        self,
        *,
        skill: dict[str, Any],
        user_text: str,
        input_files: list[dict[str, str]],
        skill_files: list[str],
    ) -> str:
        skill_name = str(skill.get("name") or skill.get("skill_id") or "skill")
        instructions = str(skill.get("skill_md_content") or skill.get("purpose") or skill.get("description") or "")
        return f"""
You are generating Python code for a generic Agent Skill runtime.

Activated Skill: {skill_name}
User task:
{user_text}

Full SKILL.md instructions:
{instructions}

Available input files copied into INPUTS_DIR:
{json.dumps(input_files, ensure_ascii=False, indent=2)}

Available files from the Skill package copied into SKILL_DIR:
{json.dumps(skill_files, ensure_ascii=False, indent=2)}

Runtime contract:
- Return JSON only, with this exact shape: {{"code": "...python script..."}}.
- The code will run in a temporary workspace. It may read files listed in INPUT_FILES and files inside SKILL_DIR.
- The code must write every final user deliverable into OUTPUTS_DIR.
- Do not write the final deliverable anywhere else.
- If editing an existing file, copy/read the referenced input file first and preserve unrelated workbook/document content as much as the Python libraries allow.
- Follow the SKILL.md literally. Use its scripts by path only when useful, for example Path(SKILL_DIR) / "scripts" / "...".
- Use formulas instead of hardcoded formula results when the Skill says so.
- Use generic reasoning from the user request and SKILL.md; do not special-case named demo prompts.
- For spreadsheet blank-column tasks, remember that a column can have a header but still be blank: decide blankness from the body/data cells, not from the header cell alone. Verify the output workbook actually removed such columns before printing success.
- At the end, print one JSON object like {{"ok": true, "message": "short human summary"}}.

Available variables in the script:
- TASK_TEXT: str
- INPUT_FILES: list of dicts with source_path, local_path, filename, extension, or error
- INPUTS_DIR: str
- OUTPUTS_DIR: str
- SKILL_DIR: str

Generate concise, complete Python code now.
""".strip()

    def _extract_script_code(self, response_text: str) -> str:
        text = (response_text or "").strip()
        if not text:
            return ""
        try:
            payload = json.loads(text)
            if isinstance(payload, dict):
                return str(payload.get("code") or "")
        except Exception:
            pass
        match = re.search(r"```(?:json)?\s*(\{[\s\S]*?\})\s*```", text, flags=re.IGNORECASE)
        if match:
            try:
                payload = json.loads(match.group(1))
                if isinstance(payload, dict):
                    return str(payload.get("code") or "")
            except Exception:
                pass
        match = re.search(r"```(?:python|py)?\s*([\s\S]*?)\s*```", text, flags=re.IGNORECASE)
        if match:
            return match.group(1).strip()
        return text if "\n" in text and ("import " in text or "from " in text) else ""

    def _validate_generated_code(self, code: str) -> None:
        try:
            tree = ast.parse(code)
        except SyntaxError as exc:
            raise ValueError(f"Syntax error: {exc}") from exc
        for node in ast.walk(tree):
            if isinstance(node, ast.Import):
                for alias in node.names:
                    root = str(alias.name or "").split(".")[0]
                    if root not in _ALLOWED_IMPORT_ROOTS:
                        raise ValueError(f"Import not allowed: {alias.name}")
            elif isinstance(node, ast.ImportFrom):
                root = str(node.module or "").split(".")[0]
                if not root or root not in _ALLOWED_IMPORT_ROOTS:
                    raise ValueError(f"Import not allowed: {node.module}")
            elif isinstance(node, ast.Call):
                name = self._call_name(node.func)
                if name in _FORBIDDEN_CALL_NAMES:
                    raise ValueError(f"Call not allowed: {name}")
                attr = self._call_attr(node.func)
                if attr in _FORBIDDEN_ATTR_NAMES:
                    raise ValueError(f"Filesystem destructive call not allowed: {attr}")

    @staticmethod
    def _call_name(node: ast.AST) -> str:
        if isinstance(node, ast.Name):
            return node.id
        return ""

    @staticmethod
    def _call_attr(node: ast.AST) -> str:
        if isinstance(node, ast.Attribute):
            return node.attr
        return ""

    def _wrap_script(self, code: str, task_path: Path) -> str:
        encoded_task_path = json.dumps(str(task_path), ensure_ascii=False)
        return f"""
from __future__ import annotations
import json
from pathlib import Path

_TASK = json.loads(Path({encoded_task_path}).read_text(encoding="utf-8"))
TASK_TEXT = _TASK.get("user_text", "")
INPUT_FILES = _TASK.get("input_files", [])
INPUTS_DIR = _TASK.get("inputs_dir", "")
OUTPUTS_DIR = _TASK.get("outputs_dir", "")
SKILL_DIR = _TASK.get("skill_dir", "")
Path(OUTPUTS_DIR).mkdir(parents=True, exist_ok=True)

{code}
""".strip() + "\n"

    def _collect_output_files(self, outputs_dir: Path) -> list[dict[str, str]]:
        generated: list[dict[str, str]] = []
        for path in outputs_dir.rglob("*"):
            if not path.is_file():
                continue
            rel = path.relative_to(outputs_dir).as_posix()
            generated.append({
                "filename": rel,
                "content_base64": base64.b64encode(path.read_bytes()).decode("ascii"),
            })
        return generated

    def _validate_generated_outputs(self, user_text: str, input_files: list[dict[str, str]], outputs_dir: Path) -> list[str]:
        issues: list[str] = []
        if not self._is_blank_column_delete_task(user_text):
            return issues

        input_file = self._first_spreadsheet_input(input_files)
        output_files = [path for path in outputs_dir.rglob("*") if path.is_file() and path.suffix.lower() in _SPREADSHEET_EXTENSIONS]
        if not input_file or not output_files:
            return issues

        try:
            source_blank_headers = self._spreadsheet_blank_headers(input_file)
        except Exception as exc:
            return [f"Unable to inspect source spreadsheet blank columns: {exc}"]
        if not source_blank_headers:
            return issues

        for output_file in output_files:
            try:
                output_headers = set(self._spreadsheet_headers(output_file))
            except Exception as exc:
                issues.append(f"Unable to inspect output spreadsheet {output_file.name}: {exc}")
                continue
            remaining = [header for header in source_blank_headers if header in output_headers]
            if remaining:
                issues.append(
                    "Output still contains blank columns that should have been removed: "
                    + ", ".join(str(item) for item in remaining)
                )
        return issues

    def _repair_outputs_for_simple_postconditions(self, user_text: str, input_files: list[dict[str, str]], outputs_dir: Path) -> str:
        if not self._is_blank_column_delete_task(user_text):
            return ""
        input_file = self._first_spreadsheet_input(input_files)
        output_files = [path for path in outputs_dir.rglob("*") if path.is_file() and path.suffix.lower() in _SPREADSHEET_EXTENSIONS]
        if not input_file or not output_files:
            return ""
        try:
            blank_headers = self._spreadsheet_blank_headers(input_file)
        except Exception:
            return ""
        if not blank_headers:
            return ""

        repaired_files: list[str] = []
        for output_file in output_files:
            try:
                from openpyxl import load_workbook

                wb = load_workbook(output_file)
                changed = False
                for ws in wb.worksheets:
                    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
                    for col in range(len(headers), 0, -1):
                        header = headers[col - 1]
                        if header in blank_headers and self._spreadsheet_column_is_blank(ws, col):
                            ws.delete_cols(col)
                            changed = True
                if changed:
                    wb.save(output_file)
                    repaired_files.append(output_file.name)
            except Exception:
                continue
        if repaired_files:
            return "Removed blank columns after postcondition validation: " + ", ".join(repaired_files)
        return ""

    def _is_blank_column_delete_task(self, user_text: str) -> bool:
        text = (user_text or "").lower()
        return any(term in text for term in _BLANK_COLUMN_TERMS) and any(term in text for term in _DELETE_TERMS)

    def _first_spreadsheet_input(self, input_files: list[dict[str, str]]) -> Path | None:
        for item in input_files:
            if item.get("error"):
                continue
            path = Path(str(item.get("local_path") or ""))
            if path.suffix.lower() in _SPREADSHEET_EXTENSIONS and path.exists():
                return path
        return None

    def _spreadsheet_headers(self, path: Path) -> list[Any]:
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=False, read_only=True)
        ws = wb.active
        return [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]

    def _spreadsheet_blank_headers(self, path: Path) -> list[Any]:
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=False, read_only=True)
        ws = wb.active
        headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        blank_headers: list[Any] = []
        for col, header in enumerate(headers, start=1):
            if header in (None, ""):
                continue
            if self._spreadsheet_column_is_blank(ws, col):
                blank_headers.append(header)
        return blank_headers

    def _spreadsheet_column_is_blank(self, ws: Any, col: int) -> bool:
        if ws.max_row < 2:
            return False
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value not in (None, ""):
                return False
        return True
    def _extract_message(self, stdout: str) -> str:
        for line in reversed((stdout or "").splitlines()):
            line = line.strip()
            if not line:
                continue
            try:
                payload = json.loads(line)
            except Exception:
                continue
            if isinstance(payload, dict) and payload.get("message"):
                return str(payload["message"])
        return ""


skill_code_runner = SkillCodeRunner()
