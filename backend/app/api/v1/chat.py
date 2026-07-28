import json
import base64
import io
import logging
import re
from datetime import datetime

from fastapi import APIRouter, Depends, Query
from sqlalchemy.orm import Session
from time import perf_counter

from app.api.deps import get_current_user_id, get_db
from app.runtime.code_executor import code_runtime_executor
from app.runtime.llm_service import LLMRequest, llm_service
from app.runtime.mcp_client import (
    extract_tool_result_text,
    get_mcp_client,
    mcp_tool_to_openai,
)
from app.runtime.knowledge_service import KnowledgeService
from app.runtime.providers import _supports_function_calling
from app.runtime.skill_executor import SkillRuntime
from app.runtime.skill_service import get_skill_package_root
from app.runtime.agent_engine import LangChainLLMWrapper, ReActAgent, ToolManager
from app.schemas.resource import (
    CodeExecutionAuditRecord,
    ChatMessageRecord,
    ChatMessageRequest,
    ChatMessageResponse,
    ChatSessionCreate,
    ChatSessionRecord,
    RuntimeRunEventRecord,
    RuntimeRunRecord,
)
from app.services.postgres_store import store
from app.services.user_file_service import user_file_service
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)
router = APIRouter()


def _is_skill_listing_query(text: str) -> bool:
    normalized = (text or "").strip().lower()
    patterns = [
        r"有哪些\s*skills?",
        r"有哪?些\s*skill",
        r"what\s+skills?\s+do\s+you\s+have",
        r"list\s+.*skills?",
        r"当前\s*skills?",
        r"有什么\s*skills?",
        r"现在\s*.*\s*什么\s*skills?",
    ]
    return any(re.search(pattern, normalized, flags=re.IGNORECASE) for pattern in patterns)


def _render_bound_skills(bound_skills: list[dict]) -> str:
    if not bound_skills:
        return "当前没有绑定任何 Skill。"

    lines = ["当前我已绑定并可用的 Skills："]
    for idx, skill in enumerate(bound_skills, start=1):
        name = skill.get("name") or skill.get("skill_id")
        version = skill.get("version") or "-"
        entrypoint = skill.get("entrypoint") or "-"
        desc = skill.get("description") or ""
        purpose = skill.get("purpose") or desc or "未提供"
        capabilities = skill.get("capabilities") or []
        caps_text = "、".join(str(item) for item in capabilities[:6]) if capabilities else "-"
        lines.append(f"{idx}. {name}")
        lines.append(f"   - 用途: {purpose}")
        lines.append(f"   - version: {version}")
        lines.append(f"   - entrypoint: {entrypoint}")
        lines.append(f"   - capabilities: {caps_text}")
        if desc:
            lines.append(f"   - description: {desc}")

    lines.append("如你希望我调用某个 Skill，请直接说“使用 <Skill 名称> 来处理 …”，我会在回复中标注使用的 Skill。")
    return "\n".join(lines)


def _extract_mentioned_skills(text: str, bound_skills: list[dict]) -> list[str]:
    normalized = (text or "").lower()
    result: list[str] = []
    for item in bound_skills:
        name = str(item.get("name") or "").strip()
        if name and name.lower() in normalized:
            result.append(name)
    return list(dict.fromkeys(result))




def _llm_answer_or_error(llm_response) -> str:
    if llm_response and llm_response.text:
        return llm_response.text
    error = getattr(llm_response, "error", None) if llm_response else "empty LLM response"
    return f"[LLM Error] {error or 'empty LLM response'}"

def _skill_display_name(skill: dict) -> str:
    return str(skill.get("name") or skill.get("skill_id") or "skill").strip()


def _skill_match_haystack(skill: dict) -> str:
    parts = [
        skill.get("name") or "",
        skill.get("description") or "",
        skill.get("purpose") or "",
        " ".join(str(item) for item in (skill.get("capabilities") or [])),
    ]
    return " ".join(str(part) for part in parts if part).lower()


def _skill_matches_request(skill: dict, user_text: str) -> bool:
    normalized = (user_text or "").strip().lower()
    if not normalized:
        return False

    name = _skill_display_name(skill).lower()
    if name and name in normalized:
        return True

    for capability in skill.get("capabilities") or []:
        capability_text = str(capability).strip().lower()
        if capability_text and capability_text in normalized:
            return True

    haystack = _skill_match_haystack(skill)
    for token in re.split(r"[\s,，。；;、/|()（）\[\]{}:：]+", normalized):
        token = token.strip()
        if len(token) >= 3 and token in haystack:
            return True

    # Chinese task phrases are often not whitespace-delimited. Use concise
    # purpose/description phrases as substring hints for activation.
    for source in (skill.get("purpose") or "", skill.get("description") or ""):
        for phrase in re.split(r"[，。；;、\n\r]+", str(source).lower()):
            phrase = phrase.strip()
            if len(phrase) >= 4 and phrase in normalized:
                return True

    return False


def _explicitly_requested_skills(bound_skills: list[dict], user_text: str) -> list[dict]:
    normalized = (user_text or "").strip().lower()
    if not normalized:
        return []

    requested: list[dict] = []
    for skill in bound_skills:
        name = _skill_display_name(skill).lower()
        if name and re.search(rf"(?<![a-z0-9_-]){re.escape(name)}(?![a-z0-9_-])", normalized):
            requested.append(skill)
    return requested


def _select_activated_skills(bound_skills: list[dict], user_text: str) -> list[dict]:
    if not bound_skills or _is_skill_listing_query(user_text):
        return []

    explicitly_requested = _explicitly_requested_skills(bound_skills, user_text)
    if explicitly_requested:
        return explicitly_requested

    matched = [skill for skill in bound_skills if _skill_matches_request(skill, user_text)]
    if matched:
        return matched

    # In the common single-skill authoring/test path, the associated skill is
    # the intended operating manual for the agent unless the user is only asking
    # for inventory.
    if len(bound_skills) == 1:
        return bound_skills

    return []


def _build_skill_runtime_prompt(bound_skills: list[dict], user_text: str) -> tuple[str, list[str]]:
    if not bound_skills:
        return "", []

    discovery_lines = [
        "Agent Skills are lightweight, reusable instruction packages. Use progressive disclosure:",
        "1. Discovery: first consider only each Skill name and short description.",
        "2. Activation: load and follow full SKILL.md instructions only when the user task matches that Skill.",
        "3. Execution: follow activated instructions strictly; use attached scripts/templates only when the instructions require them and the runtime supports that action.",
        "",
        "Available Skills (discovery view):",
    ]
    for skill in bound_skills:
        name = _skill_display_name(skill)
        purpose = skill.get("purpose") or skill.get("description") or "No short description provided"
        capabilities = skill.get("capabilities") or []
        discovery_lines.append(f"- {name}: {purpose}; capabilities={capabilities}")

    activated_skills = _select_activated_skills(bound_skills, user_text)
    activated_names = [_skill_display_name(skill) for skill in activated_skills]

    if _is_front_design_active(activated_names):
        discovery_lines.extend([
            "",
            "Artifact execution contract for frontend-design:",
            "- The primary deliverable is a downloadable .html file, not a chat explanation.",
            "- Produce one complete standalone HTML document only; the backend will save it to My Files.",
            "- Do not include markdown fences, prose, or partial snippets around the HTML document.",
        ])
    if _is_xlsx_active(activated_names):
        discovery_lines.extend([
            "",
            "Artifact execution contract for xlsx:",
            "- The primary deliverable is a spreadsheet file in My Files, not a chat-only table.",
            "- If you can emit files, return JSON with generated_files entries using filename and content/content_base64.",
            "- If code execution is available, use openpyxl/pandas to create or modify the workbook.",
        ])

    if activated_skills:
        discovery_lines.append("")
        discovery_lines.append("Activated Skill instructions for this request:")
        for skill in activated_skills:
            name = _skill_display_name(skill)
            instructions = (skill.get("skill_md_content") or skill.get("purpose") or skill.get("description") or "").strip()
            discovery_lines.append(f"\n--- Skill: {name} ---")
            discovery_lines.append(instructions or "No detailed SKILL.md instructions were uploaded for this Skill.")
        discovery_lines.append("")
        discovery_lines.append("When you use an activated Skill, include a final line exactly like: 使用的 Skill: <skill name>")
    else:
        discovery_lines.append("")
        discovery_lines.append("No full Skill instructions are activated for this request. Do not claim to use a Skill unless the user's task clearly matches one; ask a clarifying question if needed.")

    discovery_lines.append("When a user asks about available skills, answer from the discovery list only.")
    return "\n".join(discovery_lines), activated_names

def _has_skill_name(bound_skills: list[dict], skill_names: set[str]) -> bool:
    for item in bound_skills:
        name = str(item.get("name") or "").strip().lower()
        if name in skill_names:
            return True
    return False


def _looks_truncated_design_output(answer: str) -> bool:
    text = (answer or "").rstrip()
    if not text:
        return False

    lowered = text.lower()
    if text.count("```") % 2 == 1:
        return True
    if "<html" in lowered and "</html>" not in lowered:
        return True
    if "<style" in lowered and "</style>" not in lowered:
        return True

    truncated_endings = (
        "transform:",
        "background:",
        "color:",
        "padding:",
        "margin:",
        "width:",
        "height:",
        "top:",
        "left:",
        "right:",
        "bottom:",
        "opacity:",
        "font-size:",
        "line-height:",
        "{",
        ":",
        "=",
    )
    return any(text.endswith(item) for item in truncated_endings)


def _is_design_skill(bound_skills: list[dict]) -> bool:
    return _has_skill_name(bound_skills, {"frontend-design", "front-design"})


def _is_design_request(user_text: str) -> bool:
    """Check if the user's question is actually about design or page creation."""
    normalized = (user_text or "").strip().lower()
    if not normalized:
        return False

    design_keywords = {
        "设计", "design", "页面", "page", "website", "首页", "主页", "首頁",
        "前端", "frontend", "网站", "web", "homepage", "布局", "layout",
        "样式", "style", "ui", "ux", "界面", "interface", "创建", "create",
        "制作", "make", "建立", "build", "开发", "develop", "卖", "售", "花",
    }
    return any(keyword in normalized for keyword in design_keywords)


def _repair_truncated_design_output(answer: str) -> str:
    text = (answer or "").rstrip()
    if not text:
        return text

    if "\n" in text:
        text = text.rsplit("\n", 1)[0].rstrip()

    lowered = text.lower()
    suffix_parts: list[str] = []
    if "<style" in lowered and "</style>" not in lowered:
        suffix_parts.append("}")
        suffix_parts.append("</style>")
    if "<body" in lowered and "</body>" not in lowered:
        suffix_parts.append("</body>")
    if "<html" in lowered and "</html>" not in lowered:
        suffix_parts.append("</html>")
    if text.count("```") % 2 == 1:
        suffix_parts.append("```")

    if suffix_parts:
        text = f"{text}\n" + "\n".join(suffix_parts)
    return text


def _build_design_skill_brief(user_text: str, bound_skills: list[dict]) -> str:
    subject = "高端品牌网站首页"
    if user_text.strip():
        subject = user_text.strip()

    brief = [
        "Design a premium homepage for: " + subject,
        "Use the activated frontend-design SKILL.md as the operating manual, not as decorative context.",
        "Internally follow its process: brainstorm a compact design system, critique generic choices, revise, then build. Do not show the planning unless the user explicitly asks for it.",
        "Return exactly one complete single-file HTML document. Start with <!DOCTYPE html> and end with </html>. Do not use Markdown fences, do not write 'html', 'Copy', commentary, or explanations outside the file.",
        "The page must include complete <head>, <style>, and <body> sections and must be usable by saving it as an .html file.",
        "Tailor the visual language to the user's subject and intent. Infer the domain from the request, use subject-specific materials, vocabulary, interactions, and content, and avoid unrelated default luxury motifs.",
        "Make the first viewport substantial: a distinctive hero, visible brand name, strong typographic hierarchy, and a memorable signature visual built with CSS or reliable inline code-native elements. Avoid an empty page, nav-only draft, generic cards, and unfinished CSS.",
        "Use no external network dependencies for core rendering. If fonts or images fail, the page must still look intentionally designed using CSS, gradients, shapes, and system font fallbacks.",
        "The backend will save your HTML as a file, so optimize for a complete artifact rather than a conversational answer.",
    ]
    return "\n\n".join(brief)

def _skill_name_matches(names: list[str], candidates: set[str]) -> bool:
    return any(str(name).strip().lower() in candidates for name in names)


def _is_front_design_active(names: list[str]) -> bool:
    return _skill_name_matches(names, {"frontend-design", "front-design"})


def _is_xlsx_active(names: list[str]) -> bool:
    return _skill_name_matches(names, {"xlsx", "spreadsheet", "spreadsheets"})


def _strip_markdown_code_fence(text: str) -> str:
    value = (text or "").strip()
    fence = re.match(r"^```[a-zA-Z0-9_-]*\s*([\s\S]*?)\s*```$", value)
    if fence:
        return fence.group(1).strip()
    return value


def _extract_html_artifact(answer: str) -> str:
    text = _strip_markdown_code_fence(answer)
    text = re.sub(r"^\s*(html|copy|html\s+copy)\s*\n+", "", text, flags=re.IGNORECASE)
    text = text.replace("```html", "").replace("```css", "").replace("```", "")

    lowered = text.lower()
    start_match = re.search(r"<!doctype\s+html|<html[\s>]", lowered)
    if start_match:
        text = text[start_match.start():]

    lowered = text.lower()
    if "<html" not in lowered:
        return ""
    if "</html>" in lowered:
        end = lowered.rfind("</html>") + len("</html>")
        text = text[:end]
    else:
        text = _repair_truncated_design_output(text)

    return text.strip()


def _derive_artifact_title(user_text: str, fallback: str = "artifact") -> str:
    text = re.sub(r"\s+", " ", (user_text or "").strip())
    text = re.sub(r"(?i)\buse\s+[\w-]+\s+skill\b", "", text)
    text = re.sub(r"使用\s*[\w\-\u4e00-\u9fff]+\s*skill\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"(帮我|请|生成|创建|制作|设计|写一个|做一个|一个|一份|左右|即可)", "", text)
    text = re.sub(r"[，。,.!?！？；;:：].*$", "", text).strip()
    return text[:40] or fallback


def _safe_artifact_filename(title: str, suffix: str) -> str:
    safe = re.sub(r"[^\w\u4e00-\u9fff\-]+", "_", title).strip("_")
    safe = safe[:48] or "artifact"
    return f"{safe}.{suffix.lstrip('.')}"


def _html_escape(text: str) -> str:
    return (
        (text or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _build_front_design_fallback_html(user_text: str) -> str:
    title = _derive_artifact_title(user_text, fallback="品牌首页")
    escaped_title = _html_escape(title)
    escaped_brief = _html_escape((user_text or title).strip())
    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{escaped_title}</title>
<style>
:root{{--ink:#171717;--paper:#f6f2ea;--line:#d8d0c3;--accent:#0f766e;--accent2:#b45309;--deep:#202c33;--soft:#e8dfd1}}*{{box-sizing:border-box}}body{{margin:0;background:var(--paper);color:var(--ink);font-family:Inter,'Noto Sans SC',system-ui,sans-serif}}a{{color:inherit;text-decoration:none}}.nav{{height:68px;display:flex;align-items:center;justify-content:space-between;padding:0 6vw;border-bottom:1px solid var(--line);background:rgba(246,242,234,.92);backdrop-filter:blur(16px);position:sticky;top:0;z-index:10}}.brand{{font-weight:800;font-size:20px;letter-spacing:.08em}}.links{{display:flex;gap:26px;font-size:13px;color:#5c5650}}.hero{{min-height:calc(100vh - 68px);display:grid;grid-template-columns:1.05fr .95fr;gap:5vw;align-items:center;padding:64px 6vw 72px}}.kicker{{font-size:12px;text-transform:uppercase;letter-spacing:.24em;color:var(--accent);font-weight:800}}h1{{font-size:clamp(48px,7vw,96px);line-height:.95;margin:20px 0 24px;max-width:780px}}.lede{{font-size:19px;line-height:1.8;color:#56514c;max-width:680px}}.actions{{display:flex;gap:14px;flex-wrap:wrap;margin-top:34px}}.btn{{padding:14px 20px;border:1px solid var(--ink);font-size:13px;font-weight:800}}.primary{{background:var(--ink);color:var(--paper)}}.visual{{min-height:520px;position:relative;overflow:hidden;background:linear-gradient(145deg,var(--deep),#3d4d4f 58%,#a99170);border-radius:0 0 0 72px}}.plate{{position:absolute;inset:9%;border:1px solid rgba(255,255,255,.25)}}.sphere{{position:absolute;width:46%;aspect-ratio:1;border-radius:50%;background:radial-gradient(circle at 32% 28%,#fff7e8,#d4a94f 42%,#0f766e 72%);right:10%;top:11%;box-shadow:0 30px 80px rgba(0,0,0,.28)}}.bar{{position:absolute;left:11%;right:18%;height:10px;background:#f5efe5;bottom:22%;box-shadow:0 38px 0 rgba(245,239,229,.5),0 76px 0 rgba(245,239,229,.22)}}.grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:22px;padding:76px 6vw;border-top:1px solid var(--line)}}.card{{border-top:3px solid var(--accent);padding-top:22px;min-height:180px}}.card:nth-child(2){{border-color:var(--accent2)}}.card:nth-child(3){{border-color:#334155}}.card h2{{font-size:24px;margin:0 0 14px}}.card p{{line-height:1.8;color:#625d56;margin:0}}.story{{padding:84px 6vw;background:#fffaf2;border-top:1px solid var(--line)}}.story h2{{font-size:clamp(34px,5vw,64px);line-height:1.05;margin:0 0 24px;max-width:920px}}.story p{{font-size:18px;line-height:1.9;max-width:920px;color:#59534d}}footer{{padding:36px 6vw;display:flex;justify-content:space-between;gap:20px;color:#6a6258;border-top:1px solid var(--line)}}@media(max-width:860px){{.links{{display:none}}.hero{{grid-template-columns:1fr;padding-top:44px}}.visual{{min-height:360px;border-radius:0 0 0 42px}}.grid{{grid-template-columns:1fr}}footer{{display:block}}}}
</style>
</head>
<body>
<nav class="nav"><div class="brand">{escaped_title}</div><div class="links"><a href="#work">精选</a><a href="#story">叙事</a><a href="#contact">联系</a></div></nav>
<main>
<section class="hero"><div><div class="kicker">Premium homepage</div><h1>{escaped_title}</h1><p class="lede">{escaped_brief}</p><div class="actions"><a class="btn primary" href="#contact">开始咨询</a><a class="btn" href="#work">查看内容</a></div></div><div class="visual" aria-label="品牌视觉"><div class="plate"></div><div class="sphere"></div><div class="bar"></div></div></section>
<section class="grid" id="work"><article class="card"><h2>清晰主张</h2><p>把用户请求中的核心主题放到首屏，让页面一打开就能看出品牌、对象或服务。</p></article><article class="card"><h2>完整结构</h2><p>包含导航、主视觉、内容模块和行动入口，可直接保存为单文件 HTML 预览。</p></article><article class="card"><h2>可继续迭代</h2><p>样式、文案和模块都集中在文件内，便于后续按真实素材和业务细节修改。</p></article></section>
<section class="story" id="story"><h2>围绕主题建立有辨识度的首页体验。</h2><p>这是模型输出不可用时的通用兜底文件。正常情况下应优先保存由 front-design Skill 指令驱动生成的完整 HTML。</p></section>
</main>
<footer id="contact"><span>{escaped_title}</span><span>Generated with front-design Skill</span></footer>
</body>
</html>"""


def _should_create_xlsx_artifact(user_text: str) -> bool:
    normalized = (user_text or "").lower()
    return any(token in normalized for token in ("xlsx", "excel", "spreadsheet", "workbook", "表格", "电子表格", "工作簿", "表"))


def _extract_requested_row_count(user_text: str, default: int = 20) -> int:
    text = user_text or ""
    match = re.search(r"(?:大约|约|around|about)?\s*(\d{1,4})\s*(?:行|条|个|名|位|记录|rows?|items?)?", text, flags=re.IGNORECASE)
    if match:
        return max(1, min(int(match.group(1)), 500))
    if any(token in text for token in ("多一点", "多一些", "较多", "丰富一点")):
        return 25
    if any(token in text for token in ("少一点", "少一些", "简单一点")):
        return 10
    return default


def _extract_requested_columns(user_text: str) -> list[str]:
    text = user_text or ""
    match = re.search(r"(?:包含|包括|字段|列|表头)[：:\s]*(.+)", text)
    if not match:
        return []
    fragment = re.split(r"[。.;；\n]", match.group(1), maxsplit=1)[0]
    parts = [part.strip(" 　、,，和及与") for part in re.split(r"[、,，/|]+|和|及|与|(?:\s+and\s+)", fragment, flags=re.IGNORECASE)]
    columns: list[str] = []
    for part in parts:
        part = re.sub(r"(?:公式|自动计算)$", "", part).strip()
        if part and len(part) <= 24 and part not in columns:
            columns.append(part)
    return columns[:24]


def _find_active_skill(bound_skills: list[dict], active_names: list[str], candidates: set[str]) -> dict | None:
    active_set = {str(name).strip().lower() for name in active_names}
    for skill in bound_skills:
        name = _skill_display_name(skill).strip().lower()
        if name in active_set and name in candidates:
            return skill
    return None


def _run_xlsx_recalc_script(user_id: str, relative_path: str, skill: dict | None) -> dict:
    if not skill:
        return {"status": "skipped", "reason": "No activated xlsx skill package"}
    try:
        package_root = get_skill_package_root(str(skill.get("skill_id") or ""))
        script_path = package_root / "scripts" / "recalc.py"
        if not script_path.exists():
            return {"status": "skipped", "reason": "xlsx skill has no scripts/recalc.py"}

        workbook_path = user_file_service.get_file_for_download(user_id, relative_path)
        proc = subprocess.run(
            [sys.executable, str(script_path), str(workbook_path), "30"],
            cwd=str(package_root),
            text=True,
            capture_output=True,
            timeout=45,
        )
        stdout = (proc.stdout or "").strip()
        stderr = (proc.stderr or "").strip()
        parsed = None
        for line in reversed(stdout.splitlines()):
            try:
                parsed = json.loads(line.strip())
                break
            except Exception:
                continue
        payload = parsed if isinstance(parsed, dict) else {"stdout": stdout, "stderr": stderr}
        status = "succeeded" if proc.returncode == 0 and not payload.get("error") else "failed"
        return {"status": status, "returncode": proc.returncode, "payload": payload, "stderr": stderr}
    except subprocess.TimeoutExpired:
        return {"status": "failed", "error": "scripts/recalc.py timed out after 45s"}
    except Exception as exc:
        return {"status": "failed", "error": str(exc)}

def _build_generic_xlsx_artifact(user_text: str) -> bytes:
    title = _derive_artifact_title(user_text, fallback="数据表")
    columns = _extract_requested_columns(user_text)
    if not columns:
        columns = ["序号", "名称", "类别", "数值1", "数值2", "数值3", "合计", "平均", "备注"]
    elif not any(col in columns for col in ("序号", "编号", "ID", "id")):
        columns = ["序号", *columns]

    wb = Workbook()
    ws = wb.active
    ws.title = re.sub(r"[:\\/?*\[\]]+", "", title)[:31] or "Sheet1"
    last_col = len(columns)
    row_count = _extract_requested_row_count(user_text)
    first_data_row = 3
    last_data_row = first_data_row + row_count - 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=18, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="305496")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    border = Border(
        left=Side(style="thin", color="B7C9D6"),
        right=Side(style="thin", color="B7C9D6"),
        top=Side(style="thin", color="B7C9D6"),
        bottom=Side(style="thin", color="B7C9D6"),
    )
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for col, header in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color="17365D")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = min(max(len(str(header)) * 2 + 6, 12), 28)

    numeric_col_indexes = [
        idx
        for idx, col in enumerate(columns, start=1)
        if re.search(r"(数值|数量|单价|金额|价格|费用|分数|得分|score|amount|price|qty|quantity|total)", col, flags=re.IGNORECASE)
    ]
    formula_total_col = next((idx for idx, col in enumerate(columns, start=1) if col in ("合计", "总计", "总分")), None)
    formula_avg_col = next((idx for idx, col in enumerate(columns, start=1) if col in ("平均", "平均分", "均值")), None)
    quantity_col = next((idx for idx, col in enumerate(columns, start=1) if re.search(r"(数量|qty|quantity)", col, flags=re.IGNORECASE)), None)
    price_col = next((idx for idx, col in enumerate(columns, start=1) if re.search(r"(单价|价格|price)", col, flags=re.IGNORECASE)), None)
    amount_col = next((idx for idx, col in enumerate(columns, start=1) if re.search(r"(金额|amount)", col, flags=re.IGNORECASE)), None)

    for offset in range(row_count):
        row = first_data_row + offset
        for col_idx, header in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=col_idx)
            if col_idx == 1 and header in ("序号", "编号", "ID", "id"):
                cell.value = offset + 1
            elif col_idx == amount_col and quantity_col and price_col:
                cell.value = f"={get_column_letter(quantity_col)}{row}*{get_column_letter(price_col)}{row}"
            elif col_idx in numeric_col_indexes and col_idx not in (formula_total_col, formula_avg_col):
                cell.value = ((offset + 1) * (col_idx + 2)) % 100 or 100
            elif col_idx == formula_total_col:
                source_cols = [idx for idx in numeric_col_indexes if idx != col_idx and idx != formula_avg_col]
                if source_cols:
                    refs = ",".join(f"{get_column_letter(idx)}{row}" for idx in source_cols)
                    cell.value = f"=SUM({refs})"
            elif col_idx == formula_avg_col:
                source_cols = [idx for idx in numeric_col_indexes if idx != col_idx and idx != formula_total_col]
                if source_cols:
                    refs = ",".join(f"{get_column_letter(idx)}{row}" for idx in source_cols)
                    cell.value = f"=ROUND(AVERAGE({refs}),2)"
            else:
                cell.value = f"{header}{offset + 1}"
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if col_idx in numeric_col_indexes or col_idx == 1 else "left")

    summary_row = last_data_row + 2
    ws.cell(row=summary_row, column=1, value="汇总")
    for col_idx in numeric_col_indexes:
        letter = get_column_letter(col_idx)
        ws.cell(row=summary_row, column=col_idx, value=f"=SUM({letter}{first_data_row}:{letter}{last_data_row})")
    for col_idx in range(1, last_col + 1):
        cell = ws.cell(row=summary_row, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2F0D9")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(last_col)}{last_data_row}"

    meta = wb.create_sheet("README")
    meta["A1"] = "用户请求"
    meta["B1"] = user_text
    meta["A2"] = "说明"
    meta["B2"] = "此文件由 xlsx Skill 激活后的通用 Artifact 生成器创建；具体业务含义来自用户请求，不在后端固定场景。"
    meta.column_dimensions["A"].width = 16
    meta.column_dimensions["B"].width = 90

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()
def _extract_generated_files_payload(answer: str) -> list[dict]:
    text = (answer or "").strip()
    if not text:
        return []

    try:
        payload = json.loads(text)
        if isinstance(payload, dict) and isinstance(payload.get("generated_files"), list):
            return payload["generated_files"]
    except Exception:
        pass

    match = re.search(r"```json\s*(\{[\s\S]*?\})\s*```", text, flags=re.IGNORECASE)
    if match:
        try:
            payload = json.loads(match.group(1))
            if isinstance(payload, dict) and isinstance(payload.get("generated_files"), list):
                return payload["generated_files"]
        except Exception:
            return []
    return []


def _decode_code_result(code_result: dict | str) -> tuple[str, list[str], list[dict[str, str]], bool]:
    text = code_result.get("text", "") if isinstance(code_result, dict) else str(code_result)
    used_tools = list(code_result.get("used_tools", [])) if isinstance(code_result, dict) else []
    used_mcps = list(code_result.get("used_mcps", [])) if isinstance(code_result, dict) else []
    return str(text), used_tools, used_mcps, llm_service.code_requests_llm(str(text))


@router.get("/agents/{agent_id}/debug", tags=["debug"])
def debug_agent_config(
    agent_id: str,
    project_id: str = Query(...),
    user_id: str = Depends(get_current_user_id),
    db: Session = Depends(get_db),
) -> dict:
    """Debug endpoint: return agent config and associated resources."""
    agent = store.get_agent_resource_for_project(db, project_id, agent_id)
    config = dict(agent.config or {})
    mcp_ids = list(config.get("mcp_ids") or [])
    
    # Ensure mcp_ids is a list of strings
    mcp_ids = [str(mid) for mid in mcp_ids if mid]
    
    mcps = store.list_mcp_resources_for_project(db, project_id=project_id, mcp_ids=mcp_ids, actor=user_id)
    
    tools_preview: list[dict] = []
    for mcp_spec in mcps:
        mcp_name = str(mcp_spec.get("name") or "")
        try:
            client = get_mcp_client(mcp_spec)
            tools = client.list_tools()
            tools_preview.append({
                "mcp_name": mcp_name,
                "mcp_id": mcp_spec.get("id"),
                "endpoint_url": mcp_spec.get("endpoint_url"),
                "transport": mcp_spec.get("transport"),
                "tool_count": len(tools),
                "tools": [{"name": t.get("name"), "description": t.get("description", "")} for t in tools[:3]],  # first 3
            })
        except Exception as exc:
            tools_preview.append({
                "mcp_name": mcp_name,
                "mcp_id": mcp_spec.get("id"),
                "error": str(exc),
            })
    
    return {
        "agent_id": agent_id,
        "agent_name": agent.name,
        "model_provider": agent.model_provider,
        "model_name": agent.model_name,
        "run_mode": config.get("run_mode", "llm"),
        "system_prompt": config.get("system_prompt", ""),
        "mcp_ids_in_config": mcp_ids,
        "mcps_count": len(mcps),
        "mcps": tools_preview,
    }


@router.get("/code-execution-audits", response_model=list[CodeExecutionAuditRecord])
def list_code_execution_audits(
    project_id: str | None = Query(default=None),
    limit: int = Query(default=100, ge=1, le=500),
    user_id: str = Depends(get_current_user_id),
    db: Session = Depends(get_db),
) -> list[dict]:
    return store.list_code_execution_audits(db, user_id=user_id, project_id=project_id, limit=limit)


@router.post("/projects/{project_id}/sessions")
def create_chat_session(
    project_id: str,
    payload: ChatSessionCreate,
    user_id: str = Depends(get_current_user_id),
    db: Session = Depends(get_db),
) -> dict:
    return store.create_chat_session(db, project_id, user_id, payload.title)


@router.get("/projects/{project_id}/sessions", response_model=list[ChatSessionRecord])
def list_chat_sessions(
    project_id: str,
    limit: int = Query(default=50, ge=1, le=200),
    user_id: str = Depends(get_current_user_id),
    db: Session = Depends(get_db),
) -> list[dict]:
    return store.list_chat_sessions(db, project_id, user_id, limit=limit)


@router.get("/sessions/{session_id}/messages", response_model=list[ChatMessageRecord])
def list_chat_messages(
    session_id: str,
    limit: int = Query(default=200, ge=1, le=500),
    user_id: str = Depends(get_current_user_id),
    db: Session = Depends(get_db),
) -> list[dict]:
    return store.list_chat_messages_for_user(db, session_id, user_id, limit=limit)


@router.get("/sessions/{session_id}/runs", response_model=list[RuntimeRunRecord])
def list_runtime_runs(
    session_id: str,
    limit: int = Query(default=100, ge=1, le=500),
    user_id: str = Depends(get_current_user_id),
    db: Session = Depends(get_db),
) -> list[dict]:
    return store.list_runtime_runs_for_session(db, session_id, user_id, limit=limit)


@router.get("/runs/{run_id}/events", response_model=list[RuntimeRunEventRecord])
def list_runtime_run_events(
    run_id: str,
    limit: int = Query(default=500, ge=1, le=1000),
    user_id: str = Depends(get_current_user_id),
    db: Session = Depends(get_db),
) -> list[dict]:
    return store.list_runtime_run_events(db, run_id, user_id, limit=limit)


@router.post("/sessions/{session_id}/messages", response_model=ChatMessageResponse)
async def send_message(
    session_id: str,
    payload: ChatMessageRequest,
    user_id: str = Depends(get_current_user_id),
    db: Session = Depends(get_db),
) -> ChatMessageResponse:
    print(f"[send_message] ✓ Request received: session_id={session_id}, agent_id={payload.agent_id}, text={payload.text[:50]}")
    logger.info(f"[send_message] ✓ Request received: session_id={session_id}, agent_id={payload.agent_id}, text={payload.text[:50]}")
    session = store.get_chat_session_for_user(db, session_id, user_id)
    run = store.create_runtime_run(
        db=db,
        session=session,
        user_id=user_id,
        input_text=payload.text,
        agent_id=payload.agent_id,
    )
    store.append_runtime_run_event(
        db=db,
        run_id=run.id,
        stage="runtime",
        status="running",
        message="Runtime execution started",
        payload={"session_id": session.id},
    )

    model_provider: str | None = None
    model_name: str | None = None
    provider_profile: str | None = None
    provider_connection_id: str | None = None
    provider_connection: dict | None = None
    system_prompt: str | None = None
    run_mode = "llm"
    custom_code = ""
    agent_config: dict = {}
    tools: list[dict] = []
    mcps: list[dict] = []
    bound_skills: list[dict] = []
    activated_skill_names: list[str] = []
    design_skill_active = False
    executable_skill_results: list[dict] = []
    direct_skill_artifact_paths: list[str] = []
    used_knowledge_bases: list[str] = []  # initialize before agent_id block
    agent_name: str | None = None
    if payload.agent_id:
        agent_resource = store.get_agent_resource_for_project(db, session.project_id, payload.agent_id)
        agent_name = agent_resource.name
        model_provider = agent_resource.model_provider
        model_name = agent_resource.model_name
        agent_config = dict(agent_resource.config or {})
        logger.info(f"[send_message] Agent resource loaded: id={payload.agent_id}, raw_config={agent_resource.config}")
        # Override agent config with request parameters if provided
        if payload.engine_type:
            agent_config["engine_type"] = payload.engine_type
        if payload.provider_profile:
            agent_config["provider_profile"] = payload.provider_profile
        if payload.temperature is not None:
            agent_config["temperature"] = payload.temperature
        if payload.max_iterations is not None:
            agent_config["max_iterations"] = payload.max_iterations
        if payload.mcp_ids is not None:
            agent_config["mcp_ids"] = payload.mcp_ids
        provider_profile = agent_config.get("provider_profile")
        provider_connection_id = agent_config.get("provider_connection_id")
        if provider_connection_id:
            provider_connection = store.get_provider_connection_runtime_config(
                db,
                connection_id=provider_connection_id,
                actor=user_id,
            )
        system_prompt = agent_config.get("system_prompt")
        run_mode = str(agent_config.get("run_mode") or "llm").strip().lower()
        custom_code = str(agent_config.get("custom_code") or "")
        tools = store.list_tool_resources_for_project(
            db,
            project_id=session.project_id,
            tool_ids=list(agent_config.get("tool_ids") or []),
            actor=user_id,
        )
        # Ensure mcp_ids is a list of strings; convert from string if needed
        raw_mcp_ids = agent_config.get("mcp_ids") or []
        mcp_ids_list: list[str] = []
        if isinstance(raw_mcp_ids, str):
            # Handle case where mcp_ids was stored as a single string
            mcp_ids_list = [raw_mcp_ids] if raw_mcp_ids else []
        else:
            mcp_ids_list = [str(mid) for mid in raw_mcp_ids if mid]
        
        logger.info(f"[send_message] Agent {payload.agent_id} config: {agent_config}")
        logger.info(f"[send_message] Agent {payload.agent_id} has mcp_ids (raw): {raw_mcp_ids}")
        logger.info(f"[send_message] Agent {payload.agent_id} has mcp_ids (list): {mcp_ids_list}")
        
        mcps = store.list_mcp_resources_for_project(
            db,
            project_id=session.project_id,
            mcp_ids=mcp_ids_list,
            actor=user_id,
        )
        logger.info(f"[send_message] Loaded {len(mcps)} MCPs: {[m.get('name') for m in mcps]}")

        bound_skills = store.list_skill_resources_for_agent(
            db,
            project_id=session.project_id,
            agent_id=payload.agent_id,
            actor=user_id,
        )

        config_skill_ids = [str(item) for item in (agent_config.get("skill_ids") or []) if item]
        config_skills = store.list_skill_resources_for_project(
            db,
            project_id=session.project_id,
            skill_ids=config_skill_ids,
            actor=user_id,
        )

        merged: dict[str, dict] = {}
        for item in bound_skills + config_skills:
            skill_key = str(item.get("skill_id") or "")
            if not skill_key:
                continue
            if skill_key not in merged:
                merged[skill_key] = item
        bound_skills = list(merged.values())
        logger.info(
            f"[send_message] Loaded {len(bound_skills)} bound skills: "
            f"{[s.get('name') for s in bound_skills]}"
        )

        if bound_skills:
            skill_prompt, activated_skill_names = _build_skill_runtime_prompt(bound_skills, payload.text)
            design_skill_active = any(
                str(name).strip().lower() in {"frontend-design", "front-design"}
                for name in activated_skill_names
            )
            is_design_req = _is_design_request(payload.text)
            print(f"[send_message] Design skill detection: design_skill_active={design_skill_active}, is_design_request={is_design_req}, text={payload.text[:50]}")
            logger.info(f"[send_message] Design skill detection: design_skill_active={design_skill_active}, is_design_request={is_design_req}, text={payload.text[:50]}")
            if design_skill_active and is_design_req:
                print(f"[send_message] ✓ Design brief injected!")
                logger.info(f"[send_message] ✓ Design brief injected!")
                skill_prompt += "\n" + _build_design_skill_brief(payload.text, bound_skills)
            else:
                print(f"[send_message] ✗ Design brief NOT injected (active={design_skill_active}, request={is_design_req})")
                logger.info(f"[send_message] ✗ Design brief NOT injected (active={design_skill_active}, request={is_design_req})")
            if system_prompt:
                system_prompt = f"{system_prompt}\n\n{skill_prompt}"
            else:
                system_prompt = skill_prompt
            store.append_runtime_run_event(
                db=db,
                run_id=run.id,
                stage="skill",
                status="activated" if activated_skill_names else "discovered",
                message="Skill discovery and progressive disclosure applied",
                payload={"available_skills": [_skill_display_name(s) for s in bound_skills], "activated_skills": activated_skill_names},
            )

            if activated_skill_names:
                skill_runtime = SkillRuntime(db)
                activated_name_set = {str(name).strip().lower() for name in activated_skill_names}
                for skill in bound_skills:
                    skill_name = _skill_display_name(skill)
                    if skill_name.lower() not in activated_name_set:
                        continue
                    if not skill_runtime.can_execute(skill):
                        store.append_runtime_run_event(
                            db=db,
                            run_id=run.id,
                            stage="skill_runtime",
                            status="skipped",
                            message=f"Skill has no executable entrypoint: {skill_name}",
                            payload={"skill": skill_name, "entrypoint": skill.get("entrypoint") or ""},
                        )
                        continue

                    safe_runtime_skill_name = re.sub(r"[^\w\-]+", "_", skill_name)
                    runtime_output_dir = f"generated/{safe_runtime_skill_name}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
                    runtime_result = skill_runtime.execute(
                        skill=skill,
                        input_data={
                            "text": payload.text,
                            "project_id": session.project_id,
                            "session_id": session.id,
                            "agent_id": payload.agent_id,
                            "user_id": user_id,
                            "config": agent_config,
                        },
                        context={
                            "project_id": session.project_id,
                            "session_id": session.id,
                            "agent_id": payload.agent_id,
                            "user_id": user_id,
                        },
                        timeout_seconds=int((skill.get("instance_config") or {}).get("timeout_seconds") or 30),
                        user_id=user_id,
                        output_base_dir=runtime_output_dir,
                    )
                    executable_skill_results.append({"skill": skill_name, **runtime_result})
                    store.append_runtime_run_event(
                        db=db,
                        run_id=run.id,
                        stage="skill_runtime",
                        status=runtime_result.get("status") or "unknown",
                        message=f"SkillRuntime executed: {skill_name}",
                        payload={
                            "skill": skill_name,
                            "entrypoint": skill.get("entrypoint") or "",
                            "status": runtime_result.get("status"),
                            "saved_files": runtime_result.get("saved_files") or [],
                            "error": runtime_result.get("error_message"),
                            "execution_id": runtime_result.get("execution_id"),
                        },
                    )
        # RAG: Load and retrieve from knowledge bases
        knowledge_base_ids = list(agent_config.get("knowledge_base_ids") or [])
        rag_context = ""
        used_knowledge_bases = []
        
        if knowledge_base_ids:
            logger.info(f"[send_message] Agent has {len(knowledge_base_ids)} knowledge bases: {knowledge_base_ids}")
            try:
                # Generate embedding for user query (optional - text search fallback if fails)
                query_embedding = None
                try:
                    emb_result = await KnowledgeService.get_embeddings(
                        texts=[payload.text],
                        model="embedding-3",
                        embedding_provider="openai",
                    )
                    if emb_result and len(emb_result) > 0:
                        query_embedding = emb_result[0]
                except Exception as emb_err:
                    logger.warning(f"[send_message] Embedding failed, using text search: {emb_err}")

                # Build RAG context (vector search if embedding available, else text search)
                rag_context = await KnowledgeService.build_rag_context(
                    db=db,
                    agent_id=payload.agent_id,
                    query_text=payload.text,
                    query_embedding=query_embedding,
                    knowledge_base_ids=knowledge_base_ids,
                )
                
                if rag_context.strip():
                    logger.info(f"[send_message] RAG context retrieved: {len(rag_context)} chars")
                    # Inject RAG context into system prompt
                    kb_prompt = f"""
You have access to the following knowledge base content relevant to the user's query:

{rag_context}

Use this information to provide accurate and informed responses. When relevant, cite the source documents."""
                    
                    if system_prompt:
                        system_prompt = f"{system_prompt}\n\n{kb_prompt}"
                    else:
                        system_prompt = kb_prompt
                    
                    # Track which knowledge bases were used
                    used_knowledge_bases = knowledge_base_ids
                else:
                    logger.info(f"[send_message] No relevant content found in knowledge bases for this query")
            except Exception as e:
                logger.warning(f"[send_message] Failed to retrieve RAG context: {str(e)}", exc_info=True)
        
        store.append_runtime_run_event(
            db=db,
            run_id=run.id,
            stage="agent",
            status="selected",
            message="Agent selected for runtime",
            payload={
                "agent_id": payload.agent_id,
                "model_provider": model_provider,
                "model_name": model_name,
                "provider_connection_id": provider_connection_id,
                "run_mode": run_mode,
                "used_knowledge_bases": used_knowledge_bases,
            },
        )
    else:
        # No agent_id provided, apply request parameters to agent_config
        if payload.engine_type:
            agent_config["engine_type"] = payload.engine_type
        if payload.provider_profile:
            agent_config["provider_profile"] = payload.provider_profile
        if payload.temperature is not None:
            agent_config["temperature"] = payload.temperature
        if payload.max_iterations is not None:
            agent_config["max_iterations"] = payload.max_iterations
        if payload.mcp_ids is not None:
            agent_config["mcp_ids"] = payload.mcp_ids

    try:
        store.append_chat_message(db, session_id, role="user", text=payload.text, agent_id=payload.agent_id)
        used_tools: list[str] = []
        used_mcps: list[dict[str, str]] = []
        used_skills: list[str] = []

        is_skill_inventory_query = _is_skill_listing_query(payload.text)
        mentioned_skills = _extract_mentioned_skills(payload.text, bound_skills) if bound_skills else []

        if is_skill_inventory_query:
            answer = _render_bound_skills(bound_skills)
            used_skills = [str(item.get("name") or item.get("skill_id")) for item in bound_skills]
            store.append_runtime_run_event(
                db=db,
                run_id=run.id,
                stage="skill",
                status="succeeded",
                message="Returned bound skill inventory",
                payload={"skills": used_skills},
            )
        elif _is_xlsx_active(activated_skill_names) and _should_create_xlsx_artifact(payload.text):
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            xlsx_base_dir = f"generated/xlsx_{timestamp}"
            xlsx_title = _derive_artifact_title(payload.text, fallback="数据表")
            xlsx_path = f"{xlsx_base_dir}/{_safe_artifact_filename(xlsx_title, 'xlsx')}"
            user_file_service.save_bytes_file(user_id, xlsx_path, _build_generic_xlsx_artifact(payload.text))
            xlsx_recalc = _run_xlsx_recalc_script(
                user_id,
                xlsx_path,
                _find_active_skill(bound_skills, activated_skill_names, {"xlsx", "spreadsheet", "spreadsheets"}),
            )
            store.append_runtime_run_event(
                db=db,
                run_id=run.id,
                stage="skill_runtime",
                status=xlsx_recalc.get("status") or "unknown",
                message="xlsx Skill scripts/recalc.py attempted for generated workbook",
                payload={"skill": "xlsx", "file": xlsx_path, **xlsx_recalc},
            )
            answer = (
                "已使用 xlsx Skill 生成电子表格文件。\n\n"
                "文件已保存到 My Files：\n"
                f"- {xlsx_path}\n\n"
                "请到 My Files 中下载该 Excel 文件。"
            )
            direct_skill_artifact_paths.append(xlsx_path)
            used_skills = activated_skill_names
            store.append_runtime_run_event(
                db=db,
                run_id=run.id,
                stage="skill_artifact",
                status="succeeded",
                message="Generated generic xlsx artifact without LLM call",
                payload={"skill": "xlsx", "files": direct_skill_artifact_paths},
            )
        elif run_mode == "code":
            started = perf_counter()
            preview = payload.text[:200]
            store.append_runtime_run_event(
                db=db,
                run_id=run.id,
                stage="code_execution",
                status="running",
                message="Code execution started",
                payload={"input_preview": preview},
            )
            code_result = code_runtime_executor.run(
                payload.text,
                custom_code=custom_code,
                context={
                    "project_id": session.project_id,
                    "session_id": session.id,
                    "user_id": user_id,
                    "agent_id": payload.agent_id,
                    "config": agent_config,
                },
                tools=tools,
                mcps=mcps,
            )
            code_text, used_tools, used_mcps, use_llm = _decode_code_result(code_result)

            if use_llm:
                if provider_connection or model_provider or model_name:
                    llm_response = llm_service.generate(
                        LLMRequest(
                            text=payload.text,
                            model_provider=model_provider,
                            model_name=model_name,
                            provider_profile=provider_profile,
                            provider_connection_id=provider_connection_id,
                            provider_connection=provider_connection,
                            system_prompt=system_prompt,
                            max_tokens=4000 if design_skill_active else None,
                        )
                    )
                    answer = _llm_answer_or_error(llm_response)
                    store.append_runtime_run_event(
                        db=db,
                        run_id=run.id,
                        stage="llm",
                        status="succeeded" if llm_response.ok else "failed",
                        message="LLM fallback executed from code mode",
                        payload={
                            "provider": llm_response.provider,
                            "model_name": llm_response.model_name,
                            "used_fallback": llm_response.used_fallback,
                            "error": llm_response.error,
                        },
                    )
                else:
                    answer = "[code-fallback-skipped] use_llm requested but model provider/model name is not configured"
            else:
                answer = code_text

            duration_ms = int((perf_counter() - started) * 1000)
            store.append_runtime_run_event(
                db=db,
                run_id=run.id,
                stage="code_execution",
                status="succeeded",
                message="Code execution succeeded",
                payload={
                    "duration_ms": duration_ms,
                    "input_preview": preview,
                    "output_length": len(answer),
                    "used_tools": used_tools,
                    "used_mcps": used_mcps,
                },
            )
        else:
            # ------------------------------------------------------------------
            # NEW: ReAct Engine (Phase 1 - supports any OpenAI-compatible model)
            # ------------------------------------------------------------------
            engine_type = str(agent_config.get("engine_type", "legacy")).strip().lower()
            
            # Auto-enable ReAct engine if MCPs are present
            if not engine_type or engine_type == "legacy":
                mcp_ids_check = agent_config.get("mcp_ids") or []
                if mcp_ids_check and len(mcp_ids_check) > 0:
                    engine_type = "react"
                    logger.info(f"[send_message] Auto-enabling ReAct engine due to presence of MCPs")
            
            logger.info(f"[send_message] engine_type resolved to: '{engine_type}'")
            
            if engine_type == "react":
                logger.info(f"[send_message] ✓ Using ReAct Agent Engine (agent={payload.agent_id})")
                try:
                    # Initialize ReAct Agent
                    llm_wrapper = LangChainLLMWrapper(
                        llm_service=llm_service,
                        model_name=model_name or "gpt-4o-mini",
                        temperature=float(agent_config.get("temperature", 0.2)),
                        provider=model_provider or "openai",
                        provider_profile=provider_profile,
                        provider_connection_id=provider_connection_id,
                        provider_connection=provider_connection,
                    )
                    tool_manager = ToolManager()
                    agent = ReActAgent(
                        llm=llm_wrapper,
                        tool_manager=tool_manager,
                        max_iterations=int(agent_config.get("max_iterations", 10)),
                    )
                    
                    # Prepare context for agent
                    agent_context = {
                        "mcps": {m.get("id"): m for m in mcps},
                        "tools": {t.get("id"): t for t in tools},
                        "skills": {s.get("skill_id"): s for s in bound_skills},
                        "knowledge_bases": {},  # TODO: Phase 2
                    }

                    agent_config = {
                        **agent_config,
                        "skill_ids": [s.get("skill_id") for s in bound_skills],
                    }
                    
                    # Run agent
                    answer, agent_events = await agent.run(
                        user_input=payload.text,
                        agent_config=agent_config,
                        context=agent_context,
                        system_prompt=system_prompt,
                    )
                    
                    # Record agent events in RuntimeRunEvent
                    for event in agent_events:
                        store.append_runtime_run_event(
                            db=db,
                            run_id=run.id,
                            stage=f"agentic_{event['stage']}",
                            status="succeeded" if event.get("error") is None else "failed",
                            message=f"Agent step: {event['stage']} (iteration {event.get('iteration', 0)})",
                            payload=event,
                        )
                    
                    logger.info(f"[send_message] ReAct Agent completed with {len(agent_events)} events")
                    
                except Exception as e:
                    logger.error(f"[send_message] ReAct Agent failed: {str(e)}", exc_info=True)
                    answer = f"[ReAct Agent Error] {str(e)}"
                    store.append_runtime_run_event(
                        db=db,
                        run_id=run.id,
                        stage="agent",
                        status="failed",
                        message="ReAct Agent execution failed",
                        payload={"error": str(e)},
                    )
            else:
                # ------------------------------------------------------------------
                # LEGACY: Original agentic loop (function calling based)
                # ------------------------------------------------------------------
                logger.info(f"[send_message] ✗ Using LEGACY engine (engine_type={engine_type}, agent={payload.agent_id})")
                # Check if model supports function calling.
                # ------------------------------------------------------------------
                supports_fc = _supports_function_calling(model_name or "")
                logger.info(f"[send_message] Model {model_name} supports function calling: {supports_fc}")
                
                if not supports_fc and mcps:
                    # Model doesn't support function calling. Fallback to direct LLM call.
                    logger.info(f"[send_message] Model doesn't support function calling. Calling LLM without tools.")
                    llm_response = llm_service.generate(
                        LLMRequest(
                            text=payload.text,
                            model_provider=model_provider,
                            model_name=model_name,
                            provider_profile=provider_profile,
                            provider_connection_id=provider_connection_id,
                            provider_connection=provider_connection,
                            system_prompt=system_prompt,
                            messages=None,  # Use simple single-turn
                            tools=None,  # Don't send tools
                        )
                    )
                    answer = _llm_answer_or_error(llm_response)
                    store.append_runtime_run_event(
                        db=db,
                        run_id=run.id,
                        stage="llm",
                        status="succeeded" if llm_response.ok else "failed",
                        message="LLM generation completed (model doesn't support function calling)",
                        payload={
                            "provider": llm_response.provider,
                            "model_name": llm_response.model_name,
                            "note": f"MCPs available but not used: {[m.get('name') for m in mcps]}",
                        },
                    )
                else:
                    # ------------------------------------------------------------------
                    # Fetch tool definitions from associated MCPs (best-effort).
                    # ------------------------------------------------------------------
                    openai_tools: list[dict] = []
                    # Maps OpenAI tool name → (mcp_spec, original_tool_name)
                    tool_to_mcp: dict[str, tuple[dict, str]] = {}
                    mcp_names_for_prompt: list[str] = []

                    for mcp_spec in mcps:
                        mcp_name = str(mcp_spec.get("name") or "")
                        try:
                            mcp_c = get_mcp_client(mcp_spec)
                            tools_list = mcp_c.list_tools()
                            logger.info(f"[send_message] MCP {mcp_name} has {len(tools_list)} tools")
                            if tools_list:
                                mcp_names_for_prompt.append(mcp_name)
                            for tool in tools_list:
                                openai_def = mcp_tool_to_openai(mcp_name, tool)
                                fn_name = openai_def["function"]["name"]
                                openai_tools.append(openai_def)
                                tool_to_mcp[fn_name] = (mcp_spec, str(tool.get("name") or ""))
                        except Exception as exc:
                            # Log but don't fail; other MCPs may still work
                            logger.warning(f"[send_message] Failed to load MCP {mcp_name}: {exc}")

                    # ------------------------------------------------------------------
                    # Augment system prompt with available tools info.
                    # ------------------------------------------------------------------
                    final_system_prompt = system_prompt or "You are a helpful assistant."
                    if mcp_names_for_prompt:
                        tools_list_text = ", ".join(mcp_names_for_prompt)
                        augmented_prompt = f"{final_system_prompt}\n\nYou have access to the following MCP tools/services: {tools_list_text}. When appropriate, use these tools to answer user questions and get real-time information."
                    else:
                        augmented_prompt = final_system_prompt

                    # ------------------------------------------------------------------
                    # Build initial conversation messages.
                    # ------------------------------------------------------------------
                    conv_messages: list[dict] = []
                    conv_messages.append({"role": "system", "content": augmented_prompt})
                    conv_messages.append({"role": "user", "content": payload.text})

                    # ------------------------------------------------------------------
                    # Agentic loop: call LLM → execute tool calls → repeat.
                    # ------------------------------------------------------------------
                    _MAX_TOOL_ITERATIONS = 10
                    answer = ""
                    llm_response = None
                    
                    logger.info(f"[send_message] Starting agentic loop with {len(openai_tools)} tools, prompt includes: {mcp_names_for_prompt}")

                    for _iter in range(_MAX_TOOL_ITERATIONS):
                        logger.debug(f"[send_message] Agentic iteration {_iter + 1}")
                        llm_response = llm_service.generate(
                            LLMRequest(
                                text=payload.text,
                                model_provider=model_provider,
                                model_name=model_name,
                                provider_profile=provider_profile,
                                provider_connection_id=provider_connection_id,
                                provider_connection=provider_connection,
                                system_prompt=system_prompt,
                                messages=conv_messages,
                                tools=openai_tools if openai_tools else None,
                                max_tokens=4000 if design_skill_active else None,
                            )
                        )
                        
                        if llm_response.tool_calls:
                            logger.info(f"[send_message] LLM requested {len(llm_response.tool_calls)} tool calls")
                            conv_messages.append({
                                "role": "assistant",
                                "content": llm_response.text or None,
                                "tool_calls": [
                                    {
                                        "id": tc["id"],
                                        "type": "function",
                                        "function": {
                                            "name": tc["name"],
                                            "arguments": json.dumps(tc["arguments"]),
                                    },
                                }
                                for tc in llm_response.tool_calls
                            ],
                            })

                            # Execute each requested tool call.
                            for tc in llm_response.tool_calls:
                                fn_name = tc["name"]
                                fn_args = tc["arguments"]
                                logger.debug(f"[send_message] Executing tool: {fn_name}")
                                if fn_name in tool_to_mcp:
                                    mcp_spec, orig_tool = tool_to_mcp[fn_name]
                                    mcp_res_name = str(mcp_spec.get("name") or "")
                                    try:
                                        raw = get_mcp_client(mcp_spec).call_tool(orig_tool, fn_args)
                                        result_text = extract_tool_result_text(raw)
                                        used_mcps.append({"mcp": mcp_res_name, "tool": orig_tool})
                                        logger.info(f"[send_message] Tool call succeeded: {mcp_res_name}/{orig_tool}")
                                        store.append_runtime_run_event(
                                            db=db, run_id=run.id, stage="mcp", status="succeeded",
                                            message=f"MCP tool called: {mcp_res_name}/{orig_tool}",
                                            payload={"mcp": mcp_res_name, "tool": orig_tool},
                                        )
                                    except Exception as exc:
                                        result_text = f"[error calling {fn_name}: {exc}]"
                                        logger.error(f"[send_message] Tool call failed: {fn_name}, error: {exc}")
                                        store.append_runtime_run_event(
                                            db=db, run_id=run.id, stage="mcp", status="failed",
                                            message=f"MCP tool failed: {fn_name}",
                                            payload={"error": str(exc)},
                                        )
                                else:
                                    result_text = f"[unknown tool: {fn_name}]"
                                    logger.warning(f"[send_message] Unknown tool: {fn_name}")

                                conv_messages.append({
                                    "role": "tool",
                                    "tool_call_id": tc["id"],
                                    "content": result_text,
                                })
                        else:
                            answer = _llm_answer_or_error(llm_response)
                            logger.info(f"[send_message] LLM returned final answer (iteration {_iter + 1})")
                            break
                    else:
                        answer = _llm_answer_or_error(llm_response) if llm_response else "[max tool iterations reached]"

                    store.append_runtime_run_event(
                        db=db,
                        run_id=run.id,
                        stage="llm",
                        status="succeeded" if (llm_response and llm_response.ok) else "failed",
                        message="LLM generation completed",
                        payload={
                            "provider": llm_response.provider if llm_response else "",
                            "model_name": llm_response.model_name if llm_response else "",
                            "used_fallback": llm_response.used_fallback if llm_response else False,
                            "error": llm_response.error if llm_response else None,
                            "tool_iterations": _iter + 1,
                        "mcp_calls": len(used_mcps),
                    },
                )
        generated_files = _extract_generated_files_payload(answer)
        saved_file_paths: list[str] = list(direct_skill_artifact_paths)

        if design_skill_active and not answer.startswith("[runtime-fallback:"):
            for continuation_index in range(3):
                if not _looks_truncated_design_output(answer):
                    break

                continuation_prompt = (
                    "Continue the previous HTML document from the exact point it stopped. "
                    "Do not repeat earlier content. Finish open CSS, body content, scripts, and close all tags. "
                    "Output only the continuation text, with no markdown fences or commentary.\n\n"
                    f"Previous output:\n{answer}\n\nContinuation:"
                )
                continuation = llm_service.generate(
                    LLMRequest(
                        text=continuation_prompt,
                        model_provider=model_provider,
                        model_name=model_name,
                        provider_profile=provider_profile,
                        provider_connection_id=provider_connection_id,
                        provider_connection=provider_connection,
                        system_prompt=system_prompt,
                        max_tokens=2000,
                    )
                )
                if not continuation.ok or not continuation.text:
                    break

                answer = f"{answer}{continuation.text.lstrip()}"
                store.append_runtime_run_event(
                    db=db,
                    run_id=run.id,
                    stage="llm_continuation",
                    status="succeeded",
                    message=f"Continued truncated design output (pass {continuation_index + 1})",
                    payload={"continuation_length": len(continuation.text)},
                )

            if _looks_truncated_design_output(answer):
                repaired_answer = _repair_truncated_design_output(answer)
                if repaired_answer != answer:
                    answer = repaired_answer
                    store.append_runtime_run_event(
                        db=db,
                        run_id=run.id,
                        stage="llm_repair",
                        status="succeeded",
                        message="Repaired truncated design output",
                        payload={"repaired_length": len(answer)},
                    )

        if not used_skills and activated_skill_names:
            used_skills = activated_skill_names
        if not used_skills and mentioned_skills:
            used_skills = mentioned_skills

        if used_skills and not is_skill_inventory_query:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            safe_skill_name = re.sub(r"[^\w\-]+", "_", used_skills[0]) if used_skills else "skill"
            artifact_base_dir = f"generated/{safe_skill_name}_{timestamp}"
            try:
                runtime_saved_files = [
                    path
                    for result in executable_skill_results
                    for path in (result.get("saved_files") or [])
                ]
                runtime_completed = [
                    result.get("skill")
                    for result in executable_skill_results
                    if result.get("status") == "completed"
                ]
                if saved_file_paths:
                    pass
                elif runtime_saved_files:
                    saved_file_paths.extend(runtime_saved_files)
                    answer = (
                        f"已使用 {', '.join(str(item) for item in runtime_completed if item)} Skill 执行脚本并生成文件。\n\n"
                        "文件已保存到 My Files：\n"
                        + "\n".join(f"- {path}" for path in saved_file_paths)
                        + "\n\n请到 My Files 中下载。"
                    )
                elif design_skill_active:
                    html = _extract_html_artifact(answer)
                    if html and "<body" in html.lower():
                        html_path = f"{artifact_base_dir}/homepage.html"
                        user_file_service.save_text_file(user_id, html_path, html)
                        saved_file_paths.append(html_path)
                        answer = (
                            "已使用 front-design Skill 生成网站首页文件。\n\n"
                            "文件已保存到 My Files：\n"
                            f"- {html_path}\n\n"
                            "请到 My Files 中下载或预览该 HTML 文件。"
                        )
                    else:
                        answer = (
                            "front-design Skill 已激活，但模型没有返回可保存的完整 HTML 文件。"
                            "请重试，或缩小页面范围后再生成。"
                        )
                elif _is_xlsx_active(used_skills) and _should_create_xlsx_artifact(payload.text):
                    xlsx_title = _derive_artifact_title(payload.text, fallback="数据表")
                    xlsx_path = f"{artifact_base_dir}/{_safe_artifact_filename(xlsx_title, 'xlsx')}"
                    user_file_service.save_bytes_file(user_id, xlsx_path, _build_generic_xlsx_artifact(payload.text))
                    xlsx_recalc = _run_xlsx_recalc_script(
                        user_id,
                        xlsx_path,
                        _find_active_skill(bound_skills, used_skills, {"xlsx", "spreadsheet", "spreadsheets"}),
                    )
                    store.append_runtime_run_event(
                        db=db,
                        run_id=run.id,
                        stage="skill_runtime",
                        status=xlsx_recalc.get("status") or "unknown",
                        message="xlsx Skill scripts/recalc.py attempted for generated workbook",
                        payload={"skill": "xlsx", "file": xlsx_path, **xlsx_recalc},
                    )
                    saved_file_paths.append(xlsx_path)
                    answer = (
                        "已使用 xlsx Skill 生成电子表格文件。\n\n"
                        "文件已保存到 My Files：\n"
                        f"- {xlsx_path}\n\n"
                        "请到 My Files 中下载该 Excel 文件。"
                    )
                elif generated_files:
                    saved_file_paths.extend(
                        user_file_service.save_generated_files(
                            user_id=user_id,
                            base_dir=artifact_base_dir,
                            generated_files=generated_files,
                        )
                    )
                    answer = (
                        f"已使用 {', '.join(used_skills)} Skill 生成文件。\n\n"
                        "文件已保存到 My Files：\n"
                        + "\n".join(f"- {path}" for path in saved_file_paths)
                        + "\n\n请到 My Files 中下载。"
                    )
                else:
                    summary_path = f"chat_outputs/{safe_skill_name}_{timestamp}.md"
                    user_file_service.save_text_file(user_id, summary_path, answer)
                    saved_file_paths.append(summary_path)

                if saved_file_paths:
                    store.append_runtime_run_event(
                        db=db,
                        run_id=run.id,
                        stage="file_library",
                        status="succeeded",
                        message="Saved Skill artifacts to user file library",
                        payload={"files": saved_file_paths},
                    )
            except Exception as file_exc:
                store.append_runtime_run_event(
                    db=db,
                    run_id=run.id,
                    stage="file_library",
                    status="failed",
                    message="Failed to persist file library artifacts",
                    payload={"error": str(file_exc)},
                )

        design_skill_used = _is_front_design_active(used_skills)
        artifact_saved = bool(saved_file_paths) and (design_skill_used or _is_xlsx_active(used_skills) or bool(generated_files))
        if used_skills and not is_skill_inventory_query and not artifact_saved and "使用的 Skill:" not in answer:
            print(f"[send_message] Adding used_skills to answer: {used_skills}")
            answer = f"{answer}\n\n使用的 Skill: {', '.join(used_skills)}"
        else:
            print(f"[send_message] NOT adding used_skills (used_skills={used_skills}, is_inventory={is_skill_inventory_query}, artifact_saved={artifact_saved}, has_already={'使用的 Skill:' in answer})")
        store.append_chat_message(db, session_id, role="assistant", text=answer, agent_id=payload.agent_id)
        store.finish_runtime_run(
            db=db,
            run_id=run.id,
            status="succeeded",
            output_text=answer,
            error=None,
        )
        store.append_runtime_run_event(
            db=db,
            run_id=run.id,
            stage="runtime",
            status="succeeded",
            message="Runtime execution completed",
            payload={"output_length": len(answer)},
        )
        print(f"[send_message] Returning assistant response: chars={len(answer)}, used_skills={used_skills}", flush=True)
        logger.info(f"[send_message] Returning assistant response: chars={len(answer)}, used_skills={used_skills}")
        return ChatMessageResponse(
            session_id=session_id,
            role="assistant",
            agent_id=payload.agent_id,
            agent_name=agent_name,
            text=answer,
            run_id=run.id,
            used_tools=used_tools,
            used_mcps=used_mcps,
            used_knowledge_bases=used_knowledge_bases,
            used_skills=used_skills,
        )
    except Exception as exc:
        error_text = str(exc)
        if run_mode == "code":
            store.append_runtime_run_event(
                db=db,
                run_id=run.id,
                stage="code_execution",
                status="failed",
                message="Code execution failed",
                payload={
                    "error": error_text,
                    "input_preview": payload.text[:200],
                },
            )
        store.finish_runtime_run(
            db=db,
            run_id=run.id,
            status="failed",
            output_text=None,
            error=error_text,
        )
        store.append_runtime_run_event(
            db=db,
            run_id=run.id,
            stage="runtime",
            status="failed",
            message="Runtime execution failed",
            payload={"error": error_text},
        )
        raise

